from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from src.build_master_roster import (
    DEFAULT_INPUT_ROOT,
    ExtractedRow,
    SUPPORTED_EXTENSIONS,
    autosize_columns,
    clean_text,
    extract_rows_from_workbook,
    normalize_banner_id,
    normalize_status,
    style_header,
    term_sort_key,
)


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_MASTER_WORKBOOK = ROOT / "Master_FSL_Roster.xlsx"
DEFAULT_OUTPUT_WORKBOOK = ROOT / "Member_Tenure_Report.xlsx"
MASTER_REQUIRED_COLUMNS = {
    "Academic Year",
    "Term",
    "Source File",
    "Source Sheet",
    "Chapter",
    "Last Name",
    "First Name",
    "Banner ID",
    "Email",
    "Status",
    "Semester Joined",
    "Position",
}

TERMINAL_STATUSES = {
    "Graduated",
    "Alumni",
    "Suspended",
    "Revoked",
    "Resigned",
    "Transfer",
    "Inactive",
}

STATUS_PRIORITY = {
    "Graduated": 90,
    "Alumni": 85,
    "Suspended": 80,
    "Revoked": 75,
    "Resigned": 70,
    "Transfer": 65,
    "Inactive": 60,
    "Active": 50,
    "New Member": 40,
    "": 0,
}


@dataclass(frozen=True)
class MemberJourney:
    chapter: str
    last_name: str
    first_name: str
    banner_id: str
    email: str
    start_term: str
    start_basis: str
    first_new_member_term: str
    last_observed_term: str
    left_term: str
    exit_reason: str
    final_status: str
    returned_later: str
    semester_count: int
    term_history: str
    status_history: str

    def as_list(self) -> List[object]:
        return [
            self.chapter,
            self.last_name,
            self.first_name,
            self.banner_id,
            self.email,
            self.start_term,
            self.start_basis,
            self.first_new_member_term,
            self.last_observed_term,
            self.left_term,
            self.exit_reason,
            self.final_status,
            self.returned_later,
            self.semester_count,
            self.term_history,
            self.status_history,
        ]

    @property
    def confirmed_join_within_window(self) -> str:
        return "Yes" if self.first_new_member_term else "No"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build a semester-count member tenure workbook from Master_FSL_Roster.xlsx and the Copy of Rosters folder."
        )
    )
    parser.add_argument(
        "--master",
        default=str(DEFAULT_MASTER_WORKBOOK),
        help="Path to Master_FSL_Roster.xlsx. Default: Master_FSL_Roster.xlsx next to the code.",
    )
    parser.add_argument(
        "--raw-root",
        default=str(DEFAULT_INPUT_ROOT),
        help="Path to the Copy of Rosters folder. Default: Copy of Rosters next to the code.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=str(DEFAULT_OUTPUT_WORKBOOK),
        help="Output workbook path. Default: Member_Tenure_Report.xlsx",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print each raw workbook as it is processed.",
    )
    return parser.parse_args()


def row_identity(row: ExtractedRow) -> Optional[Tuple[str, ...]]:
    if row.banner_id:
        return ("banner", row.chapter.lower(), row.banner_id.lower())
    if row.email:
        return ("email", row.chapter.lower(), row.email.lower())
    if row.last_name or row.first_name:
        return ("name", row.chapter.lower(), row.last_name.lower(), row.first_name.lower())
    return None


def dedupe_term_rows(rows: Sequence[ExtractedRow]) -> List[ExtractedRow]:
    best_rows: Dict[Tuple[str, ...], ExtractedRow] = {}
    for row in rows:
        identity = row_identity(row)
        if identity is None:
            continue
        key = identity + (row.term.lower(),)
        existing = best_rows.get(key)
        if existing is None or row_score(row) > row_score(existing):
            best_rows[key] = row
    return list(best_rows.values())


def row_score(row: ExtractedRow) -> int:
    score = STATUS_PRIORITY.get(row.status, 10)
    if row.banner_id:
        score += 5
    if row.email:
        score += 3
    if row.semester_joined:
        score += 1
    return score


def choose_best_identity_row(rows: Sequence[ExtractedRow]) -> ExtractedRow:
    return max(rows, key=row_score)


def choose_status(rows: Sequence[ExtractedRow]) -> str:
    return max((row.status for row in rows), key=lambda status: STATUS_PRIORITY.get(status, 10))


def term_label_sort(term_label: str) -> Tuple[int, int, str]:
    year_match = next((part for part in term_label.split() if part.isdigit() and len(part) == 4), "9999")
    return term_sort_key(year_match, term_label)


def load_master_roster(master_path: Path) -> List[ExtractedRow]:
    rows: List[ExtractedRow] = []
    if not master_path.exists():
        return rows

    wb = load_workbook(master_path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            if ws.title.lower() == "summary":
                continue

            sheet_rows = list(ws.iter_rows(values_only=True))
            if not sheet_rows:
                continue

            header_values = [clean_text(value) for value in sheet_rows[0]]
            if not MASTER_REQUIRED_COLUMNS.issubset(set(header_values)):
                continue

            header_map = {header: idx for idx, header in enumerate(header_values)}

            for row in sheet_rows[1:]:
                chapter = get_value(row, header_map, "Chapter")
                last_name = get_value(row, header_map, "Last Name")
                first_name = get_value(row, header_map, "First Name")
                banner_id = normalize_banner_id(get_value(row, header_map, "Banner ID"))
                email = get_value(row, header_map, "Email").lower()
                status = normalize_status(get_value(row, header_map, "Status"))
                semester_joined = get_value(row, header_map, "Semester Joined")
                term = get_value(row, header_map, "Term")
                academic_year = get_value(row, header_map, "Academic Year")

                if not any([chapter, last_name, first_name, banner_id, email, status]):
                    continue

                rows.append(
                    ExtractedRow(
                        academic_year=academic_year,
                        term=term,
                        source_file=get_value(row, header_map, "Source File"),
                        source_sheet=get_value(row, header_map, "Source Sheet"),
                        chapter=chapter,
                        last_name=last_name,
                        first_name=first_name,
                        banner_id=banner_id,
                        email=email,
                        status=status,
                        semester_joined=semester_joined,
                        position=get_value(row, header_map, "Position"),
                    )
                )
    finally:
        wb.close()
    return rows


def get_value(row: Tuple[object, ...], header_map: Dict[str, int], column: str) -> str:
    idx = header_map.get(column)
    if idx is None or idx >= len(row):
        return ""
    return clean_text(row[idx])


def load_raw_rosters(raw_root: Path, verbose: bool = False) -> List[ExtractedRow]:
    rows: List[ExtractedRow] = []
    if not raw_root.exists():
        return rows

    files = sorted(path for path in raw_root.rglob("*") if path.suffix.lower() in SUPPORTED_EXTENSIONS)
    for path in files:
        extracted, _ = extract_rows_from_workbook(path, verbose=verbose)
        rows.extend(extracted)
    return rows


def build_member_journeys(rows: Sequence[ExtractedRow]) -> List[MemberJourney]:
    deduped_rows = dedupe_term_rows(rows)
    rows_by_member: Dict[Tuple[str, ...], List[ExtractedRow]] = defaultdict(list)

    for row in deduped_rows:
        identity = row_identity(row)
        if identity is None:
            continue
        rows_by_member[identity].append(row)

    journeys: List[MemberJourney] = []

    for member_rows in rows_by_member.values():
        best_identity = choose_best_identity_row(member_rows)
        rows_by_term: Dict[str, List[ExtractedRow]] = defaultdict(list)
        for row in member_rows:
            rows_by_term[row.term].append(row)

        ordered_terms = sorted(rows_by_term.keys(), key=term_label_sort)
        ordered_term_rows = [rows_by_term[term] for term in ordered_terms]
        semester_count = len(ordered_terms)

        new_member_terms = [
            term
            for term, term_rows in zip(ordered_terms, ordered_term_rows)
            if any(term_row.status == "New Member" for term_row in term_rows)
        ]

        if new_member_terms:
            start_term = new_member_terms[0]
            start_basis = "Observed New Member"
        else:
            start_term = ordered_terms[0]
            start_basis = "First Observed"

        first_new_member_term = new_member_terms[0] if new_member_terms else ""
        last_observed_term = ordered_terms[-1]
        final_status = choose_status(ordered_term_rows[-1])

        left_term = ""
        exit_reason = ""
        if final_status in TERMINAL_STATUSES:
            left_term = last_observed_term
            exit_reason = final_status

        returned_later = "No"
        for idx, term_rows in enumerate(ordered_term_rows[:-1]):
            if choose_status(term_rows) in TERMINAL_STATUSES:
                returned_later = "Yes"
                break

        term_history = " | ".join(ordered_terms)
        status_history = " | ".join(
            f"{term}: {choose_status(term_rows)}" for term, term_rows in zip(ordered_terms, ordered_term_rows)
        )

        journeys.append(
            MemberJourney(
                chapter=best_identity.chapter,
                last_name=best_identity.last_name,
                first_name=best_identity.first_name,
                banner_id=best_identity.banner_id,
                email=best_identity.email,
                start_term=start_term,
                start_basis=start_basis,
                first_new_member_term=first_new_member_term,
                last_observed_term=last_observed_term,
                left_term=left_term,
                exit_reason=exit_reason,
                final_status=final_status,
                returned_later=returned_later,
                semester_count=semester_count,
                term_history=term_history,
                status_history=status_history,
            )
        )

    return sorted(
        journeys,
        key=lambda item: (
            item.semester_count,
            item.chapter.lower(),
            item.last_name.lower(),
            item.first_name.lower(),
            item.start_term.lower(),
        ),
    )


def write_summary_sheet(wb: Workbook, journeys: Sequence[MemberJourney], master_path: Path, raw_root: Path) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    style_header(ws)

    counts_by_semester = defaultdict(int)
    confirmed_counts_by_semester = defaultdict(int)
    counts_by_exit_reason = defaultdict(int)
    returned_count = 0

    for journey in journeys:
        counts_by_semester[journey.semester_count] += 1
        if journey.first_new_member_term:
            confirmed_counts_by_semester[journey.semester_count] += 1
        if journey.exit_reason:
            counts_by_exit_reason[journey.exit_reason] += 1
        if journey.returned_later == "Yes":
            returned_count += 1

    metrics = [
        ["Master workbook", str(master_path)],
        ["Raw roster folder", str(raw_root)],
        ["Distinct member journeys", len(journeys)],
        ["Journeys with observed new-member term", sum(1 for item in journeys if item.first_new_member_term)],
        ["Journeys with inferred first-observed start", sum(1 for item in journeys if not item.first_new_member_term)],
        ["Confirmed in-window joins", sum(1 for item in journeys if item.first_new_member_term)],
        ["Unconfirmed pre-window carryovers", sum(1 for item in journeys if not item.first_new_member_term)],
        ["Members who returned after a terminal status", returned_count],
    ]

    for metric in metrics:
        ws.append(metric)

    ws.append([])
    ws.append(["Semester Count", "Member Count"])
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    for semester_count in sorted(counts_by_semester):
        ws.append([semester_count, counts_by_semester[semester_count]])

    ws.append([])
    ws.append(["Semester Count", "Confirmed In-Window Join Count"])
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    for semester_count in sorted(confirmed_counts_by_semester):
        ws.append([semester_count, confirmed_counts_by_semester[semester_count]])

    ws.append([])
    ws.append(["Exit Reason", "Member Count"])
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    if counts_by_exit_reason:
        for exit_reason in sorted(counts_by_exit_reason):
            ws.append([exit_reason, counts_by_exit_reason[exit_reason]])
    else:
        ws.append(["None observed", 0])

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_semester_sheets(wb: Workbook, journeys: Sequence[MemberJourney]) -> None:
    grouped: Dict[int, List[MemberJourney]] = defaultdict(list)
    for journey in journeys:
        grouped[journey.semester_count].append(journey)

    headers = [
        "Chapter",
        "Last Name",
        "First Name",
        "Banner ID",
        "Email",
        "Start Term",
        "Start Basis",
        "First New Member Term",
        "Last Observed Term",
        "Left Term",
        "Exit Reason",
        "Final Status",
        "Returned Later",
        "Confirmed Join In Window",
        "Semester Count",
        "Term History",
        "Status History",
    ]

    for semester_count in sorted(grouped):
        ws = wb.create_sheet(title=f"{semester_count}_Semester"[:31])
        ws.append(["All Observed Members"])
        ws[ws.max_row][0].font = Font(bold=True)
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True)
        for journey in grouped[semester_count]:
            ws.append(journey.as_list()[:13] + [journey.confirmed_join_within_window] + journey.as_list()[13:])

        ws.append([])
        ws.append(["Confirmed In-Window Joins Only"])
        ws[ws.max_row][0].font = Font(bold=True)
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True)
        for journey in grouped[semester_count]:
            if journey.first_new_member_term:
                ws.append(journey.as_list()[:13] + [journey.confirmed_join_within_window] + journey.as_list()[13:])
        ws.freeze_panes = "A2"
        autosize_columns(ws)


def build_member_tenure_report(master_path: Path, raw_root: Path, output_path: Path, verbose: bool = False) -> None:
    master_rows = load_master_roster(master_path)
    raw_rows = load_raw_rosters(raw_root, verbose=verbose)

    all_rows = master_rows + raw_rows
    if not all_rows:
        raise FileNotFoundError(
            "No usable roster rows were found in Master_FSL_Roster.xlsx or the Copy of Rosters folder."
        )

    journeys = build_member_journeys(all_rows)

    wb = Workbook()
    write_summary_sheet(wb, journeys, master_path, raw_root)
    write_semester_sheets(wb, journeys)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    master_path = Path(args.master).expanduser().resolve()
    raw_root = Path(args.raw_root).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    build_member_tenure_report(
        master_path=master_path,
        raw_root=raw_root,
        output_path=output_path,
        verbose=args.verbose,
    )
    print(f"Member tenure report created: {output_path}")


if __name__ == "__main__":
    main()
