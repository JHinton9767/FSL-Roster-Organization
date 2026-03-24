```python
from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
DEFAULT_INPUT_ROOT = ROOT / "Copy of Rosters"
SUPPORTED_EXTENSIONS = {".xls", ".xlsx", ".xlsm", ".xltx", ".xltm"}
SEMESTER_FOLDER_RE = re.compile(r"^(Fall|Spring)\s+(20\d{2})$", re.IGNORECASE)

STANDARD_COLUMNS = [
    "Academic Year",
    "Term",
    "Source File",
    "Source Sheet",
    "Chapter",
    "Last Name",
    "First Name",
    "Banner ID",
    "Email",
    "Status",
    "Semester Joined",
    "Position",
]

HEADER_ALIASES = {
    "last_name": [
        "last name",
        "lastname",
        "surname",
        "member last name",
    ],
    "first_name": [
        "first name",
        "firstname",
        "given name",
        "member first name",
    ],
    "banner_id": [
        "banner id",
        "student id",
        "banner",
        "student number",
        "banner number",
        "z number",
    ],
    "email": [
        "email",
        "e-mail",
        "email address",
        "student email",
    ],
    "status": [
        "status",
        "member status",
        "membership status",
        "roster status",
    ],
    "semester_joined": [
        "semester joined",
        "joined",
        "join term",
        "semester initiated",
        "term joined",
        "semester admitted",
        "initiation term",
    ],
    "position": [
        "position",
        "office",
        "role",
        "member/council",
        "member council",
        "title",
    ],
    "chapter": [
        "chapter",
        "organization",
        "org",
        "group",
        "fraternity/sorority",
        "fsl organization",
    ],
}

STATUS_MAP = {
    "A": "Active",
    "AL": "Alumni",
    "G": "Graduated",
    "I": "Inactive",
    "S": "Suspended",
    "N": "New Member",
    "RS": "Resigned",
    "RV": "Revoked",
    "T": "Transfer",
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def canonical_header(value: object) -> str:
    text = clean_text(value).lower()
    text = text.replace("_", " ")
    text = re.sub(r"[^a-z0-9 ]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


CANONICAL_ALIAS_MAP = {
    standard_name: {canonical_header(alias) for alias in aliases}
    for standard_name, aliases in HEADER_ALIASES.items()
}


@dataclass(frozen=True)
class ExtractedRow:
    academic_year: str
    term: str
    source_file: str
    source_sheet: str
    chapter: str
    last_name: str
    first_name: str
    banner_id: str
    email: str
    status: str
    semester_joined: str
    position: str

    def as_list(self) -> List[str]:
        return [
            self.academic_year,
            self.term,
            self.source_file,
            self.source_sheet,
            self.chapter,
            self.last_name,
            self.first_name,
            self.banner_id,
            self.email,
            self.status,
            self.semester_joined,
            self.position,
        ]


def normalize_status(value: str) -> str:
    raw = clean_text(value)
    upper = raw.upper()
    if upper in STATUS_MAP:
        return STATUS_MAP[upper]
    return raw


def normalize_banner_id(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    return re.sub(r"\.0$", "", text)


def parse_term_from_path(path: Path) -> Tuple[str, str]:
    for part in path.parts:
        match = SEMESTER_FOLDER_RE.fullmatch(part)
        if match:
            return match.group(2), f"{match.group(1).title()} {match.group(2)}"

    for candidate in [path.parent.name, path.stem]:
        match = SEMESTER_FOLDER_RE.search(candidate)
        if match:
            return match.group(2), f"{match.group(1).title()} {match.group(2)}"

    year_match = re.search(r"(20\d{2}|19\d{2})", path.stem)
    if year_match:
        return year_match.group(1), year_match.group(1)
    return "Unknown", "Unknown"


def infer_chapter(path: Path, sheet_name: str) -> str:
    ignored_names = {
        "copy of rosters",
        "rosters",
        "raw rosters",
        "master roster",
    }

    for candidate in [sheet_name, path.stem, path.parent.name]:
        cleaned = clean_text(candidate)
        if not cleaned:
            continue
        if SEMESTER_FOLDER_RE.fullmatch(cleaned):
            continue
        if cleaned.lower() in ignored_names:
            continue
        if re.fullmatch(r"(19|20)\d{2}", cleaned):
            continue
        return cleaned
    return ""


def score_header_row(values: List[object]) -> Tuple[int, Dict[str, int]]:
    matched: Dict[str, int] = {}
    canon = [canonical_header(value) for value in values]
    for idx, header in enumerate(canon):
        for standard_name, aliases in CANONICAL_ALIAS_MAP.items():
            if header in aliases and standard_name not in matched:
                matched[standard_name] = idx
    return len(matched), matched


def is_banner_row(values: List[object]) -> bool:
    text_parts = [clean_text(value) for value in values if clean_text(value)]
    if len(text_parts) != 1:
        return False

    text = text_parts[0].lower()
    return "roster" in text and any(term in text for term in ["fall", "spring", "summer", "winter"])


def find_header_row(ws) -> Tuple[Optional[int], Dict[str, int]]:
    best_score = 0
    best_row_idx = None
    best_map: Dict[str, int] = {}

    max_scan_row = min(ws.max_row, 25)
    cached_rows = list(ws.iter_rows(min_row=1, max_row=max_scan_row, values_only=True))

    for row_idx, row in enumerate(cached_rows, start=1):
        row_values = list(row)

        if is_banner_row(row_values) and row_idx < len(cached_rows):
            next_values = list(cached_rows[row_idx])
            next_score, next_map = score_header_row(next_values)
            if next_score > best_score:
                best_score = next_score
                best_row_idx = row_idx + 1
                best_map = next_map

        score, header_map = score_header_row(row_values)
        if score > best_score:
            best_score = score
            best_row_idx = row_idx
            best_map = header_map

    required = {"last_name", "first_name"}
    if best_row_idx is None or best_score < 3 or not required.issubset(best_map):
        return None, {}
    return best_row_idx, best_map


def get_cell(row: Tuple[object, ...], index: Optional[int]) -> str:
    if index is None or index >= len(row):
        return ""
    return clean_text(row[index])


def row_is_empty(values: Iterable[str]) -> bool:
    return all(not clean_text(value) for value in values)


def extract_rows_from_workbook(path: Path, verbose: bool = False) -> Tuple[List[ExtractedRow], List[str]]:
    rows: List[ExtractedRow] = []
    issues: List[str] = []

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        issues.append(f"FAILED to open {path}: {exc}")
        return rows, issues

    try:
        academic_year, term = parse_term_from_path(path)

        for ws in wb.worksheets:
            header_row_idx, header_map = find_header_row(ws)
            if header_row_idx is None:
                issues.append(f"Skipped {path.name} | sheet '{ws.title}': no usable header row found.")
                continue

            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                last_name = get_cell(row, header_map.get("last_name"))
                first_name = get_cell(row, header_map.get("first_name"))
                banner_id = normalize_banner_id(get_cell(row, header_map.get("banner_id")))
                email = get_cell(row, header_map.get("email")).lower()
                status = normalize_status(get_cell(row, header_map.get("status")))
                semester_joined = get_cell(row, header_map.get("semester_joined"))
                position = get_cell(row, header_map.get("position"))
                chapter = get_cell(row, header_map.get("chapter")) or infer_chapter(path, ws.title)

                core_values = [last_name, first_name, banner_id, email, status, semester_joined, position, chapter]
                if row_is_empty(core_values):
                    continue

                if not last_name and not first_name:
                    continue

                rows.append(
                    ExtractedRow(
                        academic_year=academic_year,
                        term=term,
                        source_file=path.name,
                        source_sheet=ws.title,
                        chapter=chapter,
                        last_name=last_name,
                        first_name=first_name,
                        banner_id=banner_id,
                        email=email,
                        status=status,
                        semester_joined=semester_joined,
                        position=position,
                    )
                )
    finally:
        wb.close()

    if verbose:
        print(f"Processed {path}")
    return rows, issues


def dedupe_rows(rows: List[ExtractedRow]) -> Tuple[List[ExtractedRow], int]:
    seen: Set[Tuple[str, ...]] = set()
    deduped: List[ExtractedRow] = []
    removed = 0

    for row in rows:
        if row.banner_id:
            key = ("banner", row.academic_year.lower(), row.term.lower(), row.banner_id.lower(), row.chapter.lower())
        elif row.email:
            key = ("email", row.academic_year.lower(), row.term.lower(), row.email.lower(), row.chapter.lower())
        else:
            key = (
                "fallback",
                row.academic_year.lower(),
                row.term.lower(),
                row.chapter.lower(),
                row.last_name.lower(),
                row.first_name.lower(),
                row.semester_joined.lower(),
            )

        if key in seen:
            removed += 1
            continue
        seen.add(key)
        deduped.append(row)

    return deduped, removed


def dedupe_same_year_banner_ids(rows: List[ExtractedRow]) -> Tuple[List[ExtractedRow], int]:
    seen: Set[Tuple[str, str]] = set()
    deduped: List[ExtractedRow] = []
    removed = 0

    for row in sorted(
        rows,
        key=lambda item: (
            item.academic_year.lower(),
            item.banner_id.lower() if item.banner_id else "zzzzzzzz",
            item.term.lower(),
            item.chapter.lower(),
            item.source_file.lower(),
            item.source_sheet.lower(),
        ),
    ):
        if row.banner_id:
            key = (row.academic_year.lower(), row.banner_id.lower())
            if key in seen:
                removed += 1
                continue
            seen.add(key)
        deduped.append(row)

    return deduped, removed


def autosize_columns(ws) -> None:
    max_widths = defaultdict(int)
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            width = len(clean_text(value))
            if width > max_widths[idx]:
                max_widths[idx] = width
    for idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 32)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def write_summary_sheet(
    wb: Workbook,
    rows: List[ExtractedRow],
    issues: List[str],
    total_files: int,
    duplicates_removed: int,
    same_year_id_removed: int,
) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    style_header(ws)

    by_year = defaultdict(int)
    by_term = defaultdict(int)
    with_banner = 0
    missing_banner = 0
    chapters = set()

    for row in rows:
        by_year[row.academic_year] += 1
        by_term[row.term] += 1
        if row.banner_id:
            with_banner += 1
        else:
            missing_banner += 1
        if row.chapter:
            chapters.add(row.chapter)

    metrics = [
        ["Input files processed", total_files],
        ["Total extracted rows", len(rows)],
        ["Rows with Banner ID", with_banner],
        ["Rows missing Banner ID", missing_banner],
        ["Distinct academic years", len(by_year)],
        ["Distinct chapters", len(chapters)],
        ["Duplicate rows removed", duplicates_removed],
        ["Same-year duplicate Banner IDs removed", same_year_id_removed],
    ]
    for item in metrics:
        ws.append(item)

    ws.append([])
    ws.append(["Academic Year", "Row Count"])
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    for academic_year in sorted(by_year.keys()):
        ws.append([academic_year, by_year[academic_year]])

    ws.append([])
    ws.append(["Term", "Row Count"])
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    for term in sorted(by_term.keys()):
        ws.append([term, by_term[term]])

    ws.append([])
    ws.append(["Import Issues"])
    ws[ws.max_row][0].font = Font(bold=True)
    if issues:
        for issue in issues:
            ws.append([issue])
    else:
        ws.append(["None"])

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_year_sheets(wb: Workbook, rows: List[ExtractedRow], chunk_size: int = 1000) -> None:
    grouped: Dict[str, List[ExtractedRow]] = defaultdict(list)
    for row in rows:
        grouped[row.academic_year].append(row)

    for academic_year in sorted(grouped.keys()):
        year_rows = sorted(
            grouped[academic_year],
            key=lambda item: (
                item.banner_id.lower() if item.banner_id else "zzzzzzzz",
                item.last_name.lower(),
                item.first_name.lower(),
                item.term.lower(),
                item.chapter.lower(),
                item.source_file.lower(),
                item.source_sheet.lower(),
            ),
        )

        for start in range(0, len(year_rows), chunk_size):
            end = min(start + chunk_size, len(year_rows))
            label_start = start + 1
            label_end = end
            sheet_name = f"{academic_year}_{label_start:04d}_{label_end:04d}"
            ws = wb.create_sheet(title=sheet_name[:31])
            ws.append(STANDARD_COLUMNS)
            style_header(ws)
            for row in year_rows[start:end]:
                ws.append(row.as_list())
            ws.freeze_panes = "A2"
            autosize_columns(ws)


def build_master_roster(
    input_root: Path,
    output_file: Path,
    chunk_size: int,
    keep_duplicates: bool,
    verbose: bool,
) -> None:
    all_rows: List[ExtractedRow] = []
    issues: List[str] = []

    files = sorted(path for path in input_root.rglob("*") if path.suffix.lower() in SUPPORTED_EXTENSIONS)
    if not files:
        raise FileNotFoundError(
            f"No Excel files found under {input_root}. Supported types: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    for path in files:
        extracted, file_issues = extract_rows_from_workbook(path, verbose=verbose)
        all_rows.extend(extracted)
        issues.extend(file_issues)

    duplicates_removed = 0
    if not keep_duplicates:
        all_rows, duplicates_removed = dedupe_rows(all_rows)

    all_rows, same_year_id_removed = dedupe_same_year_banner_ids(all_rows)

    wb = Workbook()
    write_summary_sheet(
        wb,
        all_rows,
        issues,
        total_files=len(files),
        duplicates_removed=duplicates_removed,
        same_year_id_removed=same_year_id_removed,
    )
    write_year_sheets(wb, all_rows, chunk_size=chunk_size)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb.save(output_file)
    except PermissionError:
        raise PermissionError(
            f"Cannot write to '{output_file}'. Close the workbook in Excel or use a different output filename."
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build a single FSL master roster workbook from semester folders of chapter rosters. "
            "If no input path is supplied, the script uses a local 'Copy of Rosters' folder next to the code."
        )
    )
    parser.add_argument(
        "input_root",
        nargs="?",
        default=str(DEFAULT_INPUT_ROOT),
        help="Root folder containing semester folders like 'Fall 2015' and 'Spring 2026'.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="Master_FSL_Roster.xlsx",
        help="Output workbook path. Default: Master_FSL_Roster.xlsx",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=1000,
        help="Number of rows per year sheet. Default: 1000",
    )
    parser.add_argument(
        "--keep-duplicates",
        action="store_true",
        help="Keep cross-file duplicate rows. The same-year duplicate Banner ID pass still runs.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print each workbook as it is processed.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_root = Path(args.input_root).expanduser().resolve()
    output_file = Path(args.output).expanduser().resolve()

    build_master_roster(
        input_root=input_root,
        output_file=output_file,
        chunk_size=args.chunk_size,
        keep_duplicates=args.keep_duplicates,
        verbose=args.verbose,
    )
    print(f"Master roster created: {output_file}")


if __name__ == "__main__":
    main()
```